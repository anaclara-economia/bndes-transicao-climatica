"""Gera uma camada auditável de concordância histórica para o projeto BNDES.

Este script não altera a classificação substantiva das 242 políticas. Ele
organiza a dimensão histórica já produzida, agrega o valor das bases analíticas
por chave histórica e cria uma fila de revisão priorizada pelo valor real.
Casos sem correspondência confirmada permanecem em ``Demais operações`` e não
entram no numerador de Verde estrito.
"""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
PROCESSED = ROOT / "data" / "processed"
TABLES = ROOT / "outputs" / "tables"

DIM_HIST = PROCESSED / "dim_classificacao_verde_historica.parquet"
FACT_FILES = {
    "desembolsos": PROCESSED / "desembolsos_mensais_analitico.parquet",
    "contratacoes": PROCESSED / "operacoes_bndes_analitica.parquet",
}
OUT_PARQUET = PROCESSED / "dim_concordancia_historica.parquet"
OUT_XLSX = TABLES / "concordancia_historica_bndes.xlsx"
OUT_CSV = TABLES / "fila_revisao_concordancia_historica.csv"
OUT_QA = TABLES / "concordancia_historica_qa.json"

KEYS = ["produto_norm", "instrumento_financeiro_norm"]


def aggregate_fact(path: Path, prefix: str) -> pd.DataFrame:
    columns = [
        "produto_norm",
        "instrumento_financeiro_norm",
        "valor_nominal",
        "valor_real_jun2026",
    ]
    fact = pd.read_parquet(path, columns=columns)
    fact["valor_nominal"] = pd.to_numeric(fact["valor_nominal"], errors="coerce").fillna(0.0)
    fact["valor_real_jun2026"] = pd.to_numeric(
        fact["valor_real_jun2026"], errors="coerce"
    ).fillna(0.0)
    grouped = (
        fact.groupby(KEYS, dropna=False, as_index=False)
        .agg(
            **{
                f"registros_{prefix}": ("produto_norm", "size"),
                f"valor_{prefix}_nominal": ("valor_nominal", "sum"),
                f"valor_{prefix}_real_jun2026": ("valor_real_jun2026", "sum"),
            }
        )
    )
    return grouped


def build_concordance() -> tuple[pd.DataFrame, dict[str, object]]:
    hist = pd.read_parquet(DIM_HIST).copy()
    if hist.duplicated(KEYS).any():
        duplicated = hist.loc[hist.duplicated(KEYS, keep=False), KEYS]
        raise ValueError(f"A dimensão histórica possui chaves duplicadas: {duplicated.head().to_dict('records')}")

    # A dimensão legada já contém agregados nominais. Eles são removidos antes
    # do recálculo para que a camada nova use exclusivamente os fatos analíticos
    # e não crie colunas com sufixos _x/_y.
    aggregate_columns = {
        "registros_desembolsos",
        "valor_desembolsos_nominal",
        "registros_automaticas",
        "valor_automaticas_nominal",
        "registros_nao_automaticas",
        "valor_nao_automaticas_nominal",
    }
    hist = hist.drop(columns=[c for c in aggregate_columns if c in hist.columns])
    result = hist.copy()
    for prefix, path in FACT_FILES.items():
        result = result.merge(
            aggregate_fact(path, prefix),
            on=KEYS,
            how="left",
            validate="one_to_one",
        )

    numeric_columns = [
        c for c in result.columns if c.startswith("registros_") or c.startswith("valor_")
    ]
    result[numeric_columns] = result[numeric_columns].fillna(0)
    result["registros_total"] = result[["registros_desembolsos", "registros_contratacoes"]].sum(axis=1)
    result["valor_total_nominal"] = result[
        ["valor_desembolsos_nominal", "valor_contratacoes_nominal"]
    ].sum(axis=1)
    result["valor_total_real_jun2026"] = result[
        ["valor_desembolsos_real_jun2026", "valor_contratacoes_real_jun2026"]
    ].sum(axis=1)
    result["bases_observadas"] = (
        result[["registros_desembolsos", "registros_contratacoes"]]
        .gt(0)
        .sum(axis=1)
        .map({0: "Nenhuma", 1: "Uma base", 2: "Duas bases"})
    )
    result["chave_historica"] = (
        result["produto_norm"].astype("string")
        + " | "
        + result["instrumento_financeiro_norm"].astype("string")
    )

    result["pareamento_confirmado"] = result["pareamento_confirmado"].fillna(False).astype(bool)
    result["indicador_verde_estrito"] = result["indicador_verde_estrito"].fillna(False).astype(bool)
    result["status_pareamento"] = result["status_pareamento"].fillna(
        "Sem correspondência específica"
    )
    result["regra_pareamento"] = result["regra_pareamento"].fillna(
        "Sem correspondência específica"
    )

    result = result.sort_values(
        ["pareamento_confirmado", "valor_total_real_jun2026", "produto_norm", "instrumento_financeiro_norm"],
        ascending=[True, False, True, True],
    ).reset_index(drop=True)
    result["id_concordancia"] = np.arange(1, len(result) + 1)

    nao_confirmado = ~result["pareamento_confirmado"]
    valor_nao_confirmado = float(result.loc[nao_confirmado, "valor_total_real_jun2026"].sum())
    fila = result.loc[nao_confirmado].sort_values("valor_total_real_jun2026", ascending=False).copy()
    fila["participacao_valor_nao_confirmado"] = np.where(
        valor_nao_confirmado > 0,
        fila["valor_total_real_jun2026"] / valor_nao_confirmado,
        0.0,
    )
    fila["participacao_acumulada_nao_confirmada"] = fila[
        "participacao_valor_nao_confirmado"
    ].cumsum()
    prioridade = np.select(
        [
            fila["participacao_acumulada_nao_confirmada"].le(0.80),
            fila["participacao_acumulada_nao_confirmada"].le(0.95),
        ],
        ["01 - revisar primeiro por valor", "02 - revisar depois por valor"],
        default="03 - baixa prioridade relativa",
    )
    fila["prioridade_revisao"] = prioridade
    fila["marcador_especial"] = np.where(
        fila["status_pareamento"].eq("Candidato para revisão"),
        "Candidato explícito por palavra-chave",
        "",
    )

    generic = fila["instrumento_financeiro_norm"].str.contains(
        r"OUTROS|SEM INFORMACAO|NAO INFORMADO|DEMAIS", regex=True, na=False
    )
    fila["recomendacao_revisao"] = np.where(
        fila["status_pareamento"].eq("Candidato para revisão"),
        "Validar em documento oficial; não promover ao Verde estrito sem confirmação.",
        np.where(
            generic,
            "Não automatizar; exigir evidência de linha, sublinha ou descrição do projeto.",
            "Buscar equivalência histórica em política ou circular; manter em Demais até confirmação.",
        ),
    )
    fila["decisao_manual"] = ""
    fila["fonte_evidencia"] = ""
    fila["observacao_revisor"] = ""

    result = result.merge(
        fila[
            [
                "id_concordancia",
                "participacao_valor_nao_confirmado",
                "participacao_acumulada_nao_confirmada",
                "prioridade_revisao",
                "marcador_especial",
                "recomendacao_revisao",
                "decisao_manual",
                "fonte_evidencia",
                "observacao_revisor",
            ]
        ],
        on="id_concordancia",
        how="left",
        validate="one_to_one",
    )
    result["prioridade_revisao"] = result["prioridade_revisao"].fillna("00 - equivalência confirmada")
    result["recomendacao_revisao"] = result["recomendacao_revisao"].fillna(
        "Manter equivalência atual; reproduzir na análise."
    )
    for column in ["decisao_manual", "fonte_evidencia", "observacao_revisor"]:
        result[column] = result[column].fillna("")

    coverage = {}
    for prefix, label in [("desembolsos", "Desembolsos"), ("contratacoes", "Contratações")]:
        total = float(result[f"valor_{prefix}_real_jun2026"].sum())
        confirmed = float(
            result.loc[result["pareamento_confirmado"], f"valor_{prefix}_real_jun2026"].sum()
        )
        coverage[label] = {
            "valor_total_real_jun2026": total,
            "valor_confirmado_real_jun2026": confirmed,
            "cobertura_confirmada_pct": (100 * confirmed / total) if total else 0.0,
        }

    out_columns = [
        "id_concordancia",
        "chave_historica",
        "produto",
        "instrumento_financeiro",
        "produto_norm",
        "instrumento_financeiro_norm",
        "classificacao_original",
        "classificacao_analise",
        "indicador_verde_estrito",
        "bloco_tematico",
        "status_pareamento",
        "regra_pareamento",
        "pareamento_confirmado",
        "bases_observadas",
        "registros_desembolsos",
        "valor_desembolsos_nominal",
        "valor_desembolsos_real_jun2026",
        "registros_contratacoes",
        "valor_contratacoes_nominal",
        "valor_contratacoes_real_jun2026",
        "registros_total",
        "valor_total_nominal",
        "valor_total_real_jun2026",
        "participacao_valor_nao_confirmado",
        "participacao_acumulada_nao_confirmada",
        "prioridade_revisao",
        "marcador_especial",
        "recomendacao_revisao",
        "decisao_manual",
        "fonte_evidencia",
        "observacao_revisor",
    ]
    result = result[out_columns]
    result.to_parquet(OUT_PARQUET, index=False)
    fila = result.loc[~result["pareamento_confirmado"]].sort_values(
        ["prioridade_revisao", "valor_total_real_jun2026"], ascending=[True, False]
    )
    fila.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    qa = {
        "dimensao_historica_linhas": int(len(result)),
        "chaves_unicas": bool(not result.duplicated(KEYS).any()),
        "confirmados": int(result["pareamento_confirmado"].sum()),
        "nao_confirmados": int((~result["pareamento_confirmado"]).sum()),
        "verde_estrito_historico_confirmado": int(
            (result["pareamento_confirmado"] & result["indicador_verde_estrito"]).sum()
        ),
        "cobertura_por_valor": coverage,
        "fila_revisao_linhas": int(len(fila)),
        "fontes": {key: str(path.relative_to(ROOT)) for key, path in FACT_FILES.items()},
        "saida_parquet": str(OUT_PARQUET.relative_to(ROOT)),
        "saida_excel": str(OUT_XLSX.relative_to(ROOT)),
        "saida_csv": str(OUT_CSV.relative_to(ROOT)),
    }
    OUT_QA.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    return result, qa


def style_sheet(ws, table_name: str, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)
    header_fill = PatternFill("solid", fgColor="2F3A3A")
    header_font = Font(color="FFFFFF", bold=True)
    bottom = Border(bottom=Side(style="thin", color="AAB5B2"))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bottom
    for column in range(1, ws.max_column + 1):
        values = [ws.cell(row, column).value for row in range(1, min(ws.max_row, 80) + 1)]
        width = max((len(str(value)) if value is not None else 0) for value in values) + 2
        ws.column_dimensions[get_column_letter(column)].width = min(max(width, 12), 42)
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_excel(result: pd.DataFrame, qa: dict[str, object]) -> None:
    TABLES.mkdir(parents=True, exist_ok=True)

    def excel_row(values: tuple[object, ...] | list[object]) -> list[object]:
        return [None if pd.isna(value) else value for value in values]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    n = len(result) + 1
    ws_summary.append(["Indicador", "Resultado (fórmula)", "Valor de controle", "Observação"])
    summary_rows = [
        ["Combinações históricas", f"=COUNTA('Concordancia'!$A$2:$A${n})", len(result), "Chaves únicas de produto + instrumento financeiro"],
        ["Correspondências confirmadas", f"=COUNTIF('Concordancia'!$M$2:$M${n},TRUE)", int(result["pareamento_confirmado"].sum()), "Podem alimentar o indicador confirmado"],
        ["Sem correspondência confirmada", f"=COUNTIF('Concordancia'!$M$2:$M${n},FALSE)", int((~result["pareamento_confirmado"]).sum()), "Permanecem em Demais operações"],
        ["Verde estrito confirmado", f"=COUNTIFS('Concordancia'!$M$2:$M${n},TRUE,'Concordancia'!$I$2:$I${n},TRUE)", int((result["pareamento_confirmado"] & result["indicador_verde_estrito"]).sum()), "Numerador potencial"],
        ["Cobertura por valor — desembolsos", f"=IFERROR(SUMIFS('Concordancia'!$Q$2:$Q${n},'Concordancia'!$M$2:$M${n},TRUE)/SUM('Concordancia'!$Q$2:$Q${n}),0)", qa["cobertura_por_valor"]["Desembolsos"]["cobertura_confirmada_pct"] / 100, "Participação do valor com pareamento confirmado"],
        ["Cobertura por valor — contratações", f"=IFERROR(SUMIFS('Concordancia'!$T$2:$T${n},'Concordancia'!$M$2:$M${n},TRUE)/SUM('Concordancia'!$T$2:$T${n}),0)", qa["cobertura_por_valor"]["Contratações"]["cobertura_confirmada_pct"] / 100, "Participação do valor com pareamento confirmado"],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    ws_summary.append([])
    ws_summary.append(["Regra de uso", "", "", "A classificação 30/212 não foi alterada; esta camada documenta a equivalência histórica."])
    ws_summary.append(["Regra de segurança", "", "", "Casos sem confirmação permanecem em Demais operações e não entram no Verde estrito."])

    ws_conc = wb.create_sheet("Concordancia")
    ws_conc.append(list(result.columns))
    for row in result.itertuples(index=False, name=None):
        ws_conc.append(excel_row(row))

    review = result.loc[~result["pareamento_confirmado"]].sort_values(
        ["prioridade_revisao", "valor_total_real_jun2026"], ascending=[True, False]
    )
    ws_review = wb.create_sheet("Fila_Revisao")
    ws_review.append(list(review.columns))
    for row in review.itertuples(index=False, name=None):
        ws_review.append(excel_row(row))

    ws_coverage = wb.create_sheet("Cobertura")
    ws_coverage.append(["Base", "Valor total real (jun/2026)", "Valor confirmado real (jun/2026)", "Cobertura confirmada"])
    for base in ["Desembolsos", "Contratações"]:
        vals = qa["cobertura_por_valor"][base]
        ws_coverage.append([base, vals["valor_total_real_jun2026"], vals["valor_confirmado_real_jun2026"], vals["cobertura_confirmada_pct"] / 100])

    for idx, ws in enumerate([ws_conc, ws_review, ws_coverage], start=1):
        style_sheet(ws, f"TabelaConcordancia{idx}")
    style_sheet(ws_summary, "TabelaResumo")

    for ws in [ws_conc, ws_review]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in {16, 17, 19, 20, 22, 23}:
                    cell.number_format = 'R$ #,##0.00'
                elif cell.column in {24, 25}:
                    cell.number_format = '0.00%'
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
            FormulaRule(formula=[f'$M2=TRUE'], fill=PatternFill("solid", fgColor="E5F1E8")),
        )
    for cell in ws_summary["B"] + ws_summary["C"]:
        if cell.row in {6, 7}:
            cell.number_format = "0.00%"
    for cell in ws_coverage["B"][1:] + ws_coverage["C"][1:]:
        cell.number_format = 'R$ #,##0.00'
    for cell in ws_coverage["D"][1:]:
        cell.number_format = "0.00%"

    ws_summary.freeze_panes = "A2"
    ws_summary.column_dimensions["A"].width = 42
    ws_summary.column_dimensions["B"].width = 68
    ws_summary.column_dimensions["C"].width = 20
    ws_summary.column_dimensions["D"].width = 75
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUT_XLSX)


def main() -> None:
    result, qa = build_concordance()
    write_excel(result, qa)
    print(json.dumps(qa, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
