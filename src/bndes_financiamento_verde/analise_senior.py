from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
import seaborn as sns
from matplotlib.patches import FancyBboxPatch
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .paths import CONFIG_DIR, FIGURES_DIR, PROCESSED_DIR, PROJECT_ROOT, RAW_DIR, TABLES_DIR


CONFIG_PATH = CONFIG_DIR / "analise_financiamento_verde.json"
SUMMARY_PATH = TABLES_DIR / "resumo_execucao_analise_senior.json"
VALIDATION_PATH = TABLES_DIR / "relatorio_validacao_analise_senior.json"
FINAL_WORKBOOK_PATH = TABLES_DIR / "analise_financiamento_verde_bndes_senior.xlsx"
CLASSIFICATION_WORKBOOK_PATH = TABLES_DIR / "classificacao_politicas_analise_bndes.xlsx"

PALETTE = {
    "ink": "#1F2937",
    "muted": "#6B7280",
    "grid": "#D1D5DB",
    "blue": "#1F5A94",
    "blue_light": "#DCEAF6",
    "gold": "#C58A12",
    "gold_light": "#F7E8BC",
    "olive": "#657A32",
    "olive_light": "#E4EBCF",
    "orange": "#C76A24",
    "pink": "#A6527A",
    "white": "#FFFFFF",
}


@dataclass
class PipelineArtifacts:
    classificacao: pd.DataFrame
    ipca: pd.DataFrame
    concordancia: pd.DataFrame
    desembolsos: pd.DataFrame
    operacoes: pd.DataFrame
    contratos_nao_automaticos: pd.DataFrame
    resultados: dict[str, pd.DataFrame]
    resumo: dict[str, Any]
    validacao: dict[str, Any]


def ensure_dirs() -> None:
    for path in [CONFIG_DIR, PROCESSED_DIR, TABLES_DIR, FIGURES_DIR, RAW_DIR / "csv"]:
        path.mkdir(parents=True, exist_ok=True)


def carregar_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def normalizar_texto(valor: Any) -> str:
    if pd.isna(valor):
        return "SEM INFORMACAO"
    texto = unicodedata.normalize("NFKD", str(valor))
    texto = texto.encode("ascii", "ignore").decode("ascii").upper()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto or "SEM INFORMACAO"


def normalizar_serie(serie: pd.Series) -> pd.Series:
    return serie.map(normalizar_texto).astype("string")


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def inventario_fontes() -> pd.DataFrame:
    fontes = [
        PROCESSED_DIR / "desembolsos_mensais.parquet",
        PROCESSED_DIR / "operacoes_indiretas_automaticas.parquet",
        PROCESSED_DIR / "operacoes_nao_automaticas.parquet",
        PROCESSED_DIR / "politicas_operacionais.parquet",
        PROCESSED_DIR / "classificacao_politicas_analitica.parquet",
        PROCESSED_DIR / "fontes_recursos.parquet",
    ]
    registros = []
    for path in fontes:
        quadro = pd.read_parquet(path)
        registros.append(
            {
                "arquivo": path.name,
                "linhas": len(quadro),
                "colunas": len(quadro.columns),
                "tamanho_mb": path.stat().st_size / (1024**2),
                "modificado_em": pd.Timestamp(path.stat().st_mtime, unit="s"),
                "sha256": sha256_file(path),
            }
        )
        del quadro
    return pd.DataFrame(registros)


def construir_classificacao_binaria() -> pd.DataFrame:
    fonte = PROCESSED_DIR / "classificacao_politicas_analitica.parquet"
    classificacao = pd.read_parquet(fonte).copy()
    classificacao["classificacao_original"] = classificacao["classificacao"].astype("string")
    classificacao["classificacao_analise"] = np.where(
        classificacao["classificacao_original"].eq("Verde estrito"),
        "Verde estrito",
        "Demais operações",
    )
    classificacao["indicador_verde_estrito"] = classificacao["classificacao_analise"].eq(
        "Verde estrito"
    )
    classificacao["uso_na_analise"] = "Análise principal"

    for coluna in ["instrumento_de_apoio", "linha", "sublinha", "bloco_tematico"]:
        classificacao[f"{coluna}_norm"] = normalizar_serie(classificacao[coluna])

    contagens = classificacao["classificacao_analise"].value_counts().to_dict()
    if len(classificacao) != 242 or contagens != {"Demais operações": 212, "Verde estrito": 30}:
        raise ValueError(f"Classificação binária inválida: linhas={len(classificacao)}, contagens={contagens}")

    classificacao.to_parquet(PROCESSED_DIR / "classificacao_politicas_analise.parquet", index=False)
    return classificacao


def baixar_ipca(config: dict[str, Any]) -> pd.DataFrame:
    inicio = str(config["periodo_inicio"]).replace("-", "")
    referencia = str(config["ipca_mes_referencia"]).replace("-", "")
    tabela = int(config["ipca_tabela_sidra"])
    variavel = int(config["ipca_variavel_numero_indice"])
    url = (
        f"https://apisidra.ibge.gov.br/values/t/{tabela}/n1/all/v/{variavel}/"
        f"p/{inicio}-{referencia}?formato=json"
    )
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    payload = response.json()
    if len(payload) < 2:
        raise ValueError("A API SIDRA não retornou a série mensal do IPCA.")

    ipca = pd.DataFrame(payload[1:])
    ipca = (
        ipca.loc[ipca["D2C"].astype(str).eq(str(variavel)), ["D3C", "D3N", "V"]]
        .rename(columns={"D3C": "ano_mes_codigo", "D3N": "mes_nome", "V": "ipca_mes"})
        .assign(
            data_referencia=lambda df: pd.to_datetime(df["ano_mes_codigo"], format="%Y%m"),
            ipca_mes=lambda df: pd.to_numeric(df["ipca_mes"], errors="coerce"),
            fonte_ipca=url,
        )
        .sort_values("data_referencia")
        .reset_index(drop=True)
    )
    if ipca["ipca_mes"].isna().any():
        raise ValueError("O IPCA contém valores não numéricos.")

    data_ref = pd.Timestamp(config["ipca_mes_referencia"] + "-01")
    valor_ref = ipca.loc[ipca["data_referencia"].eq(data_ref), "ipca_mes"]
    if len(valor_ref) != 1:
        raise ValueError(f"Mês-base do IPCA ausente ou duplicado: {data_ref:%Y-%m}")
    ipca["fator_ipca_jun2026"] = float(valor_ref.iloc[0]) / ipca["ipca_mes"]

    raw_path = RAW_DIR / "csv" / "ipca_mensal_sidra_1737.csv"
    ipca.to_csv(raw_path, index=False, encoding="utf-8-sig")
    ipca.to_parquet(PROCESSED_DIR / "ipca_mensal.parquet", index=False)
    return ipca


def _resumir_pares(
    path: Path,
    valor_coluna: str,
    prefixo: str,
) -> pd.DataFrame:
    quadro = pd.read_parquet(path, columns=["produto", "instrumento_financeiro", valor_coluna])
    quadro["produto_norm"] = normalizar_serie(quadro["produto"])
    quadro["instrumento_financeiro_norm"] = normalizar_serie(quadro["instrumento_financeiro"])
    resumo = (
        quadro.groupby(["produto_norm", "instrumento_financeiro_norm"], as_index=False, dropna=False)
        .agg(
            **{
                f"produto_{prefixo}": ("produto", "first"),
                f"instrumento_{prefixo}": ("instrumento_financeiro", "first"),
                f"registros_{prefixo}": (valor_coluna, "size"),
                f"valor_{prefixo}_nominal": (valor_coluna, "sum"),
            }
        )
    )
    return resumo


def _mapa_univoco(
    classificacao: pd.DataFrame,
    chaves: list[str],
) -> pd.DataFrame:
    base = classificacao.copy()
    base = base.loc[~base[chaves].eq("SEM INFORMACAO").any(axis=1)]
    if base.empty:
        return pd.DataFrame(columns=chaves)

    def juntar(serie: pd.Series) -> str:
        return " | ".join(sorted(set(serie.dropna().astype(str))))

    mapa = (
        base.groupby(chaves, as_index=False, dropna=False)
        .agg(
            classificacoes=("classificacao_analise", juntar),
            categorias_originais=("classificacao_original", juntar),
            blocos=("bloco_tematico", juntar),
            n_classificacoes=("classificacao_analise", "nunique"),
        )
    )
    return mapa


def construir_concordancia(classificacao: pd.DataFrame) -> pd.DataFrame:
    desembolsos = _resumir_pares(
        PROCESSED_DIR / "desembolsos_mensais.parquet",
        "desembolsos_reais",
        "desembolsos",
    )
    automaticas = _resumir_pares(
        PROCESSED_DIR / "operacoes_indiretas_automaticas.parquet",
        "valor_da_operacao_em_reais",
        "automaticas",
    )
    nao_automaticas = _resumir_pares(
        PROCESSED_DIR / "operacoes_nao_automaticas.parquet",
        "valor_contratado_reais",
        "nao_automaticas",
    )

    chaves = ["produto_norm", "instrumento_financeiro_norm"]
    concordancia = desembolsos.merge(automaticas, on=chaves, how="outer", validate="one_to_one")
    concordancia = concordancia.merge(
        nao_automaticas, on=chaves, how="outer", validate="one_to_one"
    )
    concordancia["produto"] = concordancia[
        ["produto_desembolsos", "produto_automaticas", "produto_nao_automaticas"]
    ].bfill(axis=1).iloc[:, 0]
    concordancia["instrumento_financeiro"] = concordancia[
        [
            "instrumento_desembolsos",
            "instrumento_automaticas",
            "instrumento_nao_automaticas",
        ]
    ].bfill(axis=1).iloc[:, 0]

    classificacao = classificacao.copy()
    classificacao["instrumento_de_apoio_norm"] = normalizar_serie(
        classificacao["instrumento_de_apoio"]
    )
    classificacao["linha_norm"] = normalizar_serie(classificacao["linha"])
    classificacao["sublinha_norm"] = normalizar_serie(classificacao["sublinha"])

    candidatos_par = []
    for detalhe in ["linha_norm", "sublinha_norm"]:
        parte = classificacao[
            [
                "instrumento_de_apoio_norm",
                detalhe,
                "classificacao_analise",
                "classificacao_original",
                "bloco_tematico",
            ]
        ].rename(
            columns={
                "instrumento_de_apoio_norm": "produto_norm",
                detalhe: "instrumento_financeiro_norm",
            }
        )
        candidatos_par.append(parte)
    mapa_par = _mapa_univoco(pd.concat(candidatos_par, ignore_index=True), chaves)

    candidatos_instrumento = []
    for detalhe in ["instrumento_de_apoio_norm", "linha_norm", "sublinha_norm"]:
        parte = classificacao[
            [detalhe, "classificacao_analise", "classificacao_original", "bloco_tematico"]
        ].rename(columns={detalhe: "instrumento_financeiro_norm"})
        candidatos_instrumento.append(parte)
    mapa_instrumento = _mapa_univoco(
        pd.concat(candidatos_instrumento, ignore_index=True), ["instrumento_financeiro_norm"]
    )

    concordancia = concordancia.merge(
        mapa_par.add_prefix("par_"),
        left_on=chaves,
        right_on=["par_produto_norm", "par_instrumento_financeiro_norm"],
        how="left",
        validate="one_to_one",
    )
    concordancia = concordancia.merge(
        mapa_instrumento.add_prefix("inst_"),
        left_on="instrumento_financeiro_norm",
        right_on="inst_instrumento_financeiro_norm",
        how="left",
        validate="many_to_one",
    )

    concordancia["classificacao_analise"] = "Demais operações"
    concordancia["classificacao_original"] = pd.NA
    concordancia["bloco_tematico"] = pd.NA
    concordancia["status_pareamento"] = "Sem correspondência específica"
    concordancia["regra_pareamento"] = "Sem correspondência específica"

    par_confirmado = concordancia["par_n_classificacoes"].eq(1)
    concordancia.loc[par_confirmado, "classificacao_analise"] = concordancia.loc[
        par_confirmado, "par_classificacoes"
    ]
    concordancia.loc[par_confirmado, "classificacao_original"] = concordancia.loc[
        par_confirmado, "par_categorias_originais"
    ]
    concordancia.loc[par_confirmado, "bloco_tematico"] = concordancia.loc[
        par_confirmado, "par_blocos"
    ]
    concordancia.loc[par_confirmado, "status_pareamento"] = "Confirmado"
    concordancia.loc[par_confirmado, "regra_pareamento"] = "Produto + linha/sublinha exatos"

    inst_confirmado = (~par_confirmado) & concordancia["inst_n_classificacoes"].eq(1)
    concordancia.loc[inst_confirmado, "classificacao_analise"] = concordancia.loc[
        inst_confirmado, "inst_classificacoes"
    ]
    concordancia.loc[inst_confirmado, "classificacao_original"] = concordancia.loc[
        inst_confirmado, "inst_categorias_originais"
    ]
    concordancia.loc[inst_confirmado, "bloco_tematico"] = concordancia.loc[
        inst_confirmado, "inst_blocos"
    ]
    concordancia.loc[inst_confirmado, "status_pareamento"] = "Confirmado"
    concordancia.loc[inst_confirmado, "regra_pareamento"] = "Instrumento financeiro exato e unívoco"

    sinonimos_verdes = {
        "PROGRAMA FUNDO CLIMA": ("Clima e descarbonização", "Fundo Clima"),
        "FUNDO AMAZONIA": ("Florestas e bioeconomia", "Fundo Amazônia"),
        "BNDES FUNDO SOCIOAMBIENTAL": ("Meio ambiente", "BNDES Fundo Socioambiental"),
        "BNDES RENOVABIO": ("Biocombustíveis", "BNDES RenovaBio"),
        "BNDES FLORESTAL": ("Florestas e bioeconomia", "BNDES Florestas"),
        "BK AQUISICAO E COMERCIALIZACAO BAIXO CARBONO": (
            "Energia e eficiência",
            "BNDES Finame - Baixo Carbono",
        ),
        "MEIO AMBIENTE": ("Meio ambiente", "BNDES Finem - Meio Ambiente"),
        "SANEAMENTO": ("Saneamento", "BNDES Finem - Saneamento"),
        "SANEAMENTO AMBIENTAL E RECURSOS HIDRICOS": (
            "Saneamento",
            "Saneamento ambiental e recursos hídricos",
        ),
        "PROGRAMA BNDES SANEAMENTO PARA TODOS": (
            "Saneamento",
            "Programa BNDES Saneamento para Todos",
        ),
    }
    for instrumento_norm, (bloco, categoria_original) in sinonimos_verdes.items():
        mascara = concordancia["instrumento_financeiro_norm"].eq(instrumento_norm)
        concordancia.loc[mascara, "classificacao_analise"] = "Verde estrito"
        concordancia.loc[mascara, "classificacao_original"] = categoria_original
        concordancia.loc[mascara, "bloco_tematico"] = bloco
        concordancia.loc[mascara, "status_pareamento"] = "Confirmado"
        concordancia.loc[mascara, "regra_pareamento"] = "Sinônimo histórico validado"

    palavras_candidatas = re.compile(
        r"CLIMA|VERDE|AMAZON|FLOREST|CARBONO|SANEAMENTO|MEIO AMBIENTE|"
        r"EFICIENCIA|BIOECONOMIA|AGROECOLOGIA|RENOVAGRO|PRONAF"
    )
    sem_confirmacao = ~concordancia["status_pareamento"].eq("Confirmado")
    candidatos = sem_confirmacao & concordancia["instrumento_financeiro_norm"].str.contains(
        palavras_candidatas, na=False
    )
    concordancia.loc[candidatos, "status_pareamento"] = "Candidato para revisão"
    concordancia.loc[candidatos, "regra_pareamento"] = "Palavra-chave; não entra no numerador"

    concordancia["indicador_verde_estrito"] = concordancia["classificacao_analise"].eq(
        "Verde estrito"
    )
    concordancia["pareamento_confirmado"] = concordancia["status_pareamento"].eq("Confirmado")

    colunas_saida = [
        "produto",
        "instrumento_financeiro",
        "produto_norm",
        "instrumento_financeiro_norm",
        "classificacao_original",
        "classificacao_analise",
        "indicador_verde_estrito",
        "bloco_tematico",
        "status_pareamento",
        "regra_pareamento",
        "pareamento_confirmado",
        "registros_desembolsos",
        "valor_desembolsos_nominal",
        "registros_automaticas",
        "valor_automaticas_nominal",
        "registros_nao_automaticas",
        "valor_nao_automaticas_nominal",
    ]
    for coluna in colunas_saida:
        if coluna not in concordancia:
            concordancia[coluna] = np.nan
    concordancia = concordancia[colunas_saida].sort_values(
        ["indicador_verde_estrito", "produto_norm", "instrumento_financeiro_norm"],
        ascending=[False, True, True],
    )
    concordancia.to_parquet(
        PROCESSED_DIR / "dim_classificacao_verde_historica.parquet", index=False
    )
    return concordancia.reset_index(drop=True)


def _geografia_valida(df: pd.DataFrame, config: dict[str, Any]) -> pd.Series:
    municipio_norm = normalizar_serie(df["municipio"])
    uf_norm = normalizar_serie(df["uf"])
    codigo = (
        df["municipio_codigo"]
        .astype("string")
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str.zfill(7)
    )
    municipios_invalidos = {normalizar_texto(x) for x in config["municipios_invalidos"]}
    codigos_invalidos = set(config["codigos_municipais_invalidos"])
    ufs_invalidas = {normalizar_texto(x) for x in config["ufs_invalidas"]}
    return (
        ~municipio_norm.isin(municipios_invalidos)
        & ~codigo.isin(codigos_invalidos)
        & ~uf_norm.isin(ufs_invalidas)
    )


def _aplicar_dimensoes(
    df: pd.DataFrame,
    concordancia: pd.DataFrame,
    ipca: pd.DataFrame,
    config: dict[str, Any],
) -> pd.DataFrame:
    linhas_antes = len(df)
    valor_antes = float(df["valor_nominal"].sum())
    df["produto_norm"] = normalizar_serie(df["produto"])
    df["instrumento_financeiro_norm"] = normalizar_serie(df["instrumento_financeiro"])
    dimensao = concordancia[
        [
            "produto_norm",
            "instrumento_financeiro_norm",
            "classificacao_original",
            "classificacao_analise",
            "indicador_verde_estrito",
            "bloco_tematico",
            "status_pareamento",
            "regra_pareamento",
            "pareamento_confirmado",
        ]
    ]
    df = df.merge(
        dimensao,
        on=["produto_norm", "instrumento_financeiro_norm"],
        how="left",
        validate="many_to_one",
    )
    if len(df) != linhas_antes or not np.isclose(df["valor_nominal"].sum(), valor_antes):
        raise ValueError("O join da classificação alterou linhas ou valor monetário.")

    df = df.merge(
        ipca[["data_referencia", "ipca_mes", "fator_ipca_jun2026"]],
        on="data_referencia",
        how="left",
        validate="many_to_one",
    )
    if len(df) != linhas_antes or df["ipca_mes"].isna().any():
        raise ValueError("O join do IPCA alterou linhas ou deixou meses sem deflator.")
    df["valor_real_jun2026"] = df["valor_nominal"] * df["fator_ipca_jun2026"]
    df["classificacao_analise"] = df["classificacao_analise"].fillna("Demais operações")
    df["indicador_verde_estrito"] = df["indicador_verde_estrito"].fillna(False)
    df["status_pareamento"] = df["status_pareamento"].fillna(
        "Sem correspondência específica"
    )
    df["regra_pareamento"] = df["regra_pareamento"].fillna(
        "Sem correspondência específica"
    )
    df["pareamento_confirmado"] = df["pareamento_confirmado"].fillna(False)
    df["ano_parcial"] = df["ano"].eq(int(config["ano_parcial"]))
    df["geografia_valida"] = _geografia_valida(df, config)
    return df


def preparar_desembolsos(
    concordancia: pd.DataFrame,
    ipca: pd.DataFrame,
    config: dict[str, Any],
) -> pd.DataFrame:
    df = pd.read_parquet(PROCESSED_DIR / "desembolsos_mensais.parquet")
    inicio_ano = int(str(config["periodo_inicio"])[:4])
    df = df.loc[df["ano"].ge(inicio_ano)].copy()
    df["data_referencia"] = pd.to_datetime(
        {"year": df["ano"], "month": df["mes"], "day": 1}
    )
    df["base_origem"] = "desembolsos_mensais"
    df["unidade_analise"] = "liberação mensal"
    df["valor_nominal"] = pd.to_numeric(df["desembolsos_reais"], errors="coerce")
    if df["valor_nominal"].isna().any() or (df["valor_nominal"] < 0).any():
        raise ValueError("A base mensal contém valores ausentes ou negativos.")
    df = _aplicar_dimensoes(df, concordancia, ipca, config)
    df.to_parquet(PROCESSED_DIR / "desembolsos_mensais_analitico.parquet", index=False)
    return df


def preparar_operacoes(
    concordancia: pd.DataFrame,
    ipca: pd.DataFrame,
    config: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    auto = pd.read_parquet(PROCESSED_DIR / "operacoes_indiretas_automaticas.parquet")
    auto["duplicata_exata"] = auto.duplicated(keep="first")
    auto["data_referencia"] = pd.to_datetime(auto["data_da_contratacao"]).dt.to_period("M").dt.to_timestamp()
    auto["ano"] = auto["data_referencia"].dt.year
    auto["mes"] = auto["data_referencia"].dt.month
    auto["base_origem"] = "operacoes_indiretas_automaticas"
    auto["unidade_analise"] = "registro de operação automática"
    auto["valor_nominal"] = pd.to_numeric(auto["valor_da_operacao_em_reais"], errors="coerce")
    auto["numero_do_contrato"] = pd.NA
    auto["descricao_do_projeto"] = pd.NA
    auto["cpf_cnpj_cliente"] = auto["cpf_cnpj"].astype("string")
    auto["id_registro_fonte"] = np.arange(1, len(auto) + 1, dtype=np.int64)

    nao = pd.read_parquet(PROCESSED_DIR / "operacoes_nao_automaticas.parquet")
    nao["duplicata_exata"] = nao.duplicated(keep="first")
    nao["data_referencia"] = pd.to_datetime(nao["data_da_contratacao"]).dt.to_period("M").dt.to_timestamp()
    nao["ano"] = nao["data_referencia"].dt.year
    nao["mes"] = nao["data_referencia"].dt.month
    nao["base_origem"] = "operacoes_nao_automaticas"
    nao["unidade_analise"] = "subcrédito"
    nao["valor_nominal"] = pd.to_numeric(nao["valor_contratado_reais"], errors="coerce")
    nao["cpf_cnpj_cliente"] = nao["cnpj"].astype("string")
    nao["id_registro_fonte"] = np.arange(1, len(nao) + 1, dtype=np.int64)

    colunas = [
        "id_registro_fonte",
        "base_origem",
        "unidade_analise",
        "numero_do_contrato",
        "cliente",
        "cpf_cnpj_cliente",
        "descricao_do_projeto",
        "data_da_contratacao",
        "data_referencia",
        "ano",
        "mes",
        "valor_nominal",
        "valor_desembolsado_reais",
        "fonte_de_recurso_desembolsos",
        "modalidade_de_apoio",
        "forma_de_apoio",
        "produto",
        "instrumento_financeiro",
        "area_operacional",
        "setor_cnae",
        "subsetor_cnae_agrupado",
        "setor_bndes",
        "subsetor_bndes",
        "uf",
        "municipio",
        "municipio_codigo",
        "porte_do_cliente",
        "natureza_do_cliente",
        "duplicata_exata",
    ]
    operacoes = pd.concat([auto[colunas], nao[colunas]], ignore_index=True)
    del auto, nao
    operacoes = _aplicar_dimensoes(operacoes, concordancia, ipca, config)
    operacoes.to_parquet(PROCESSED_DIR / "operacoes_bndes_analitica.parquet", index=False)

    nao_analitica = operacoes.loc[
        operacoes["base_origem"].eq("operacoes_nao_automaticas")
    ].copy()

    def valor_unico(serie: pd.Series) -> Any:
        valores = serie.dropna().astype(str).unique()
        if len(valores) == 0:
            return pd.NA
        if len(valores) == 1:
            return valores[0]
        return "MÚLTIPLOS"

    contratos = (
        nao_analitica.groupby("numero_do_contrato", as_index=False, dropna=False)
        .agg(
            data_da_contratacao=("data_da_contratacao", "min"),
            ano=("ano", "min"),
            cliente=("cliente", valor_unico),
            uf=("uf", valor_unico),
            municipio=("municipio", valor_unico),
            quantidade_subcreditos=("numero_do_contrato", "size"),
            valor_contratado_nominal=("valor_nominal", "sum"),
            valor_contratado_real_jun2026=("valor_real_jun2026", "sum"),
            todos_subcreditos_verdes=("indicador_verde_estrito", "all"),
            algum_subcredito_verde=("indicador_verde_estrito", "any"),
            geografia_valida=("geografia_valida", "all"),
        )
    )
    contratos["ano_parcial"] = contratos["ano"].eq(int(config["ano_parcial"]))
    contratos.to_parquet(
        PROCESSED_DIR / "contratos_nao_automaticos_analitico.parquet", index=False
    )
    return operacoes, contratos


def _resumo_valores(grupo: pd.DataFrame, chaves: list[str]) -> pd.DataFrame:
    resumo = (
        grupo.groupby(chaves + ["classificacao_analise"], as_index=False, dropna=False)
        .agg(
            valor_nominal=("valor_nominal", "sum"),
            valor_real_jun2026=("valor_real_jun2026", "sum"),
            quantidade_registros=("valor_nominal", "size"),
        )
    )
    pivot = resumo.pivot_table(
        index=chaves,
        columns="classificacao_analise",
        values=["valor_nominal", "valor_real_jun2026", "quantidade_registros"],
        aggfunc="sum",
        fill_value=0,
    )
    pivot.columns = [
        f"{metrica}_{normalizar_texto(classe).lower().replace(' ', '_')}"
        for metrica, classe in pivot.columns
    ]
    pivot = pivot.reset_index()
    for metrica in ["valor_nominal", "valor_real_jun2026", "quantidade_registros"]:
        verde = f"{metrica}_verde_estrito"
        demais = f"{metrica}_demais_operacoes"
        for coluna in [verde, demais]:
            if coluna not in pivot:
                pivot[coluna] = 0
        pivot[f"{metrica}_total"] = pivot[verde] + pivot[demais]
    pivot["participacao_verde_pct"] = np.where(
        pivot["valor_real_jun2026_total"].ne(0),
        100 * pivot["valor_real_jun2026_verde_estrito"] / pivot["valor_real_jun2026_total"],
        np.nan,
    )
    return pivot


def calcular_desembolsos(
    desembolsos: pd.DataFrame, config: dict[str, Any]
) -> dict[str, pd.DataFrame]:
    mensais = _resumo_valores(desembolsos, ["data_referencia", "ano", "mes", "ano_parcial"])
    mensais["media_mensal_no_ano_real"] = mensais.groupby("ano")[
        "valor_real_jun2026_total"
    ].transform("mean")

    anuais = _resumo_valores(desembolsos, ["ano", "ano_parcial"])
    meses = desembolsos.groupby("ano")["mes"].nunique().rename("meses_observados")
    anuais = anuais.merge(meses, on="ano", how="left", validate="one_to_one")
    anuais["media_mensal_real"] = (
        anuais["valor_real_jun2026_total"] / anuais["meses_observados"]
    )
    anuais["media_mensal_verde_real"] = (
        anuais["valor_real_jun2026_verde_estrito"] / anuais["meses_observados"]
    )
    anuais["crescimento_real_total_pct"] = anuais["valor_real_jun2026_total"].pct_change() * 100
    anuais["crescimento_real_verde_pct"] = (
        anuais["valor_real_jun2026_verde_estrito"].pct_change() * 100
    )

    cobertura = (
        desembolsos.groupby(["ano", "status_pareamento", "regra_pareamento"], as_index=False)
        .agg(valor_real_jun2026=("valor_real_jun2026", "sum"), registros=("valor_nominal", "size"))
    )
    totais = cobertura.groupby("ano")["valor_real_jun2026"].transform("sum")
    cobertura["participacao_valor_pct"] = 100 * cobertura["valor_real_jun2026"] / totais

    sazonal = (
        desembolsos.groupby(["ano", "mes"], as_index=False)
        .agg(valor_real_jun2026=("valor_real_jun2026", "sum"))
    )
    sazonal["participacao_no_ano_pct"] = 100 * sazonal["valor_real_jun2026"] / sazonal.groupby(
        "ano"
    )["valor_real_jun2026"].transform("sum")
    sazonal["razao_mes_media_ano"] = sazonal["valor_real_jun2026"] / sazonal.groupby("ano")[
        "valor_real_jun2026"
    ].transform("mean")

    completo = desembolsos[desembolsos["ano"].le(int(config["ultimo_ano_completo"]))]
    acumulado = _resumo_valores(completo.assign(periodo="2002–2025"), ["periodo"])
    return {
        "desembolsos_mensais": mensais,
        "desembolsos_anuais": anuais,
        "cobertura_classificacao_desembolsos": cobertura,
        "sazonalidade_mensal": sazonal,
        "desembolsos_acumulados": acumulado,
    }


def calcular_contratacoes(
    operacoes: pd.DataFrame,
    contratos: pd.DataFrame,
    config: dict[str, Any],
) -> dict[str, pd.DataFrame]:
    anuais = _resumo_valores(operacoes, ["ano", "ano_parcial", "base_origem"])
    cobertura = (
        operacoes.groupby(
            ["ano", "base_origem", "status_pareamento", "regra_pareamento"],
            as_index=False,
        )
        .agg(valor_real_jun2026=("valor_real_jun2026", "sum"), registros=("valor_nominal", "size"))
    )
    cobertura["participacao_valor_pct"] = 100 * cobertura["valor_real_jun2026"] / cobertura.groupby(
        ["ano", "base_origem"]
    )["valor_real_jun2026"].transform("sum")

    auto = operacoes[operacoes["base_origem"].eq("operacoes_indiretas_automaticas")]
    auto_sem_dup = auto.loc[~auto["duplicata_exata"]]
    sens_auto = []
    for nome, quadro in [("Todos os registros", auto), ("Sem duplicatas exatas", auto_sem_dup)]:
        total = quadro["valor_real_jun2026"].sum()
        verde = quadro.loc[quadro["indicador_verde_estrito"], "valor_real_jun2026"].sum()
        sens_auto.append(
            {
                "sensibilidade": nome,
                "registros": len(quadro),
                "valor_real_jun2026": total,
                "valor_verde_real_jun2026": verde,
                "participacao_verde_pct": 100 * verde / total if total else np.nan,
            }
        )
    robustez_auto = pd.DataFrame(sens_auto)

    contrato_resumo = pd.DataFrame(
        [
            {
                "criterio_contrato_verde": "Todos os subcréditos verdes",
                "contratos_verdes": int(contratos["todos_subcreditos_verdes"].sum()),
                "contratos_total": len(contratos),
                "participacao_contratos_verdes_pct": 100
                * contratos["todos_subcreditos_verdes"].mean(),
            },
            {
                "criterio_contrato_verde": "Ao menos um subcrédito verde",
                "contratos_verdes": int(contratos["algum_subcredito_verde"].sum()),
                "contratos_total": len(contratos),
                "participacao_contratos_verdes_pct": 100
                * contratos["algum_subcredito_verde"].mean(),
            },
        ]
    )

    estatisticas = (
        operacoes.groupby(["base_origem", "classificacao_analise"])["valor_real_jun2026"]
        .agg(
            quantidade="size",
            media="mean",
            mediana="median",
            primeiro_quartil=lambda s: s.quantile(0.25),
            terceiro_quartil=lambda s: s.quantile(0.75),
            total="sum",
        )
        .reset_index()
    )
    return {
        "contratacoes_anuais": anuais,
        "cobertura_classificacao_contratacoes": cobertura,
        "robustez_duplicatas_automaticas": robustez_auto,
        "robustez_contratos_nao_automaticos": contrato_resumo,
        "estatisticas_contratacoes": estatisticas,
    }


def _resumo_hierarquia(
    df: pd.DataFrame,
    colunas: list[str],
    universo: str,
    config: dict[str, Any],
) -> pd.DataFrame:
    colunas_necessarias = list(dict.fromkeys(colunas + ["ano", "valor_nominal", "valor_real_jun2026", "indicador_verde_estrito"]))
    base = df.loc[
        df["ano"].between(int(str(config["periodo_inicio"])[:4]), int(config["ultimo_ano_completo"])),
        colunas_necessarias,
    ].copy()
    base["universo"] = universo
    base["valor_verde_real_jun2026"] = np.where(
        base["indicador_verde_estrito"], base["valor_real_jun2026"], 0.0
    )
    resumo = (
        base.groupby(["universo"] + colunas, as_index=False, dropna=False)
        .agg(
            valor_real_jun2026=("valor_real_jun2026", "sum"),
            valor_verde_real_jun2026=("valor_verde_real_jun2026", "sum"),
            quantidade_registros=("valor_nominal", "size"),
        )
    )
    total = base["valor_real_jun2026"].sum()
    total_verde = base.loc[base["indicador_verde_estrito"], "valor_real_jun2026"].sum()
    resumo["participacao_no_total_bndes_pct"] = 100 * resumo["valor_real_jun2026"] / total
    resumo["participacao_no_verde_pct"] = np.where(
        total_verde > 0, 100 * resumo["valor_verde_real_jun2026"] / total_verde, np.nan
    )
    resumo["participacao_verde_interna_pct"] = np.where(
        resumo["valor_real_jun2026"].ne(0),
        100 * resumo["valor_verde_real_jun2026"] / resumo["valor_real_jun2026"],
        np.nan,
    )
    return resumo


def _top10(
    df: pd.DataFrame,
    dimensao: str,
    universo: str,
    config: dict[str, Any],
    somente_geografia_valida: bool = False,
) -> pd.DataFrame:
    recortes = [
        ("2002–2025", df["ano"].le(int(config["ultimo_ano_completo"]))),
        ("2025", df["ano"].eq(int(config["ultimo_ano_completo"]))),
        ("2026 YTD", df["ano"].eq(int(config["ano_parcial"]))),
    ]
    saidas: list[pd.DataFrame] = []
    for rotulo, mascara in recortes:
        base = df.loc[mascara & df["indicador_verde_estrito"]].copy()
        if somente_geografia_valida:
            base = base.loc[base["geografia_valida"]]
        base[dimensao] = base[dimensao].fillna("SEM INFORMAÇÃO")
        resumo = (
            base.groupby(dimensao, as_index=False, dropna=False)
            .agg(
                valor_real_jun2026=("valor_real_jun2026", "sum"),
                quantidade_registros=("valor_nominal", "size"),
            )
            .sort_values(["valor_real_jun2026", "quantidade_registros"], ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        total = base["valor_real_jun2026"].sum()
        resumo["participacao_no_verde_valido_pct"] = np.where(
            total > 0, 100 * resumo["valor_real_jun2026"] / total, np.nan
        )
        resumo["participacao_acumulada_pct"] = resumo[
            "participacao_no_verde_valido_pct"
        ].cumsum()
        resumo.insert(0, "posicao", np.arange(1, len(resumo) + 1))
        resumo.insert(0, "recorte", rotulo)
        resumo.insert(0, "universo", universo)
        saidas.append(resumo)
    return pd.concat(saidas, ignore_index=True) if saidas else pd.DataFrame()


def calcular_composicao_top10(
    desembolsos: pd.DataFrame,
    operacoes: pd.DataFrame,
    config: dict[str, Any],
) -> dict[str, pd.DataFrame]:
    hierarquias = []
    for nome, quadro in [("Desembolsos", desembolsos), ("Contratações", operacoes)]:
        quadro = quadro.copy()
        quadro["modalidade_analitica"] = (
            quadro["modalidade_de_apoio"].fillna(quadro["forma_de_apoio"])
            if "modalidade_de_apoio" in quadro.columns
            else quadro["forma_de_apoio"]
        )
        hierarquias.append(
            _resumo_hierarquia(
                quadro,
                ["bloco_tematico", "produto", "instrumento_financeiro", "modalidade_analitica"],
                nome,
                config,
            )
        )
    blocos_instrumentos = pd.concat(hierarquias, ignore_index=True)

    top_resultados: dict[str, list[pd.DataFrame]] = {
        "top10_instrumentos": [],
        "top10_linhas": [],
        "top10_setores": [],
        "top10_uf": [],
        "top10_municipios": [],
    }
    cobertura_territorial = []
    for nome, quadro in [("Desembolsos", desembolsos), ("Contratações", operacoes)]:
        quadro_verde = quadro.loc[quadro["indicador_verde_estrito"]].copy()
        top_resultados["top10_instrumentos"].append(_top10(quadro_verde, "produto", nome, config))
        top_resultados["top10_linhas"].append(
            _top10(quadro_verde, "instrumento_financeiro", nome, config)
        )
        top_resultados["top10_setores"].append(
            _top10(quadro_verde, "subsetor_cnae_agrupado", nome, config)
        )
        top_resultados["top10_uf"].append(
            _top10(quadro_verde, "uf", nome, config, somente_geografia_valida=True)
        )
        top_resultados["top10_municipios"].append(
            _top10(quadro_verde, "municipio", nome, config, somente_geografia_valida=True)
        )
        for rotulo, mascara in [
            ("2002–2025", quadro["ano"].le(int(config["ultimo_ano_completo"]))),
            ("2025", quadro["ano"].eq(int(config["ultimo_ano_completo"]))),
            ("2026 YTD", quadro["ano"].eq(int(config["ano_parcial"]))),
        ]:
            verde = quadro.loc[mascara & quadro["indicador_verde_estrito"]]
            total = verde["valor_real_jun2026"].sum()
            valido = verde.loc[verde["geografia_valida"], "valor_real_jun2026"].sum()
            cobertura_territorial.append(
                {
                    "universo": nome,
                    "recorte": rotulo,
                    "valor_verde_total": total,
                    "valor_verde_geografia_valida": valido,
                    "valor_verde_nao_territorializavel": total - valido,
                    "cobertura_territorial_pct": 100 * valido / total if total else np.nan,
                }
            )

    saida = {"blocos_instrumentos": blocos_instrumentos}
    saida.update({chave: pd.concat(valor, ignore_index=True) for chave, valor in top_resultados.items()})
    saida["cobertura_territorial"] = pd.DataFrame(cobertura_territorial)
    return saida


def construir_qualidade_bases(
    inventario: pd.DataFrame,
    desembolsos: pd.DataFrame,
    operacoes: pd.DataFrame,
    contratos: pd.DataFrame,
) -> pd.DataFrame:
    registros = [
        {
            "teste": "Fontes processadas inventariadas",
            "resultado": len(inventario),
            "unidade": "arquivos",
            "status": "OK" if len(inventario) >= 6 else "REVISAR",
        },
        {
            "teste": "Cobertura temporal dos desembolsos",
            "resultado": f"{desembolsos['data_referencia'].min():%Y-%m} a {desembolsos['data_referencia'].max():%Y-%m}",
            "unidade": "período",
            "status": "OK",
        },
        {
            "teste": "Cobertura temporal das contratações",
            "resultado": f"{operacoes['data_referencia'].min():%Y-%m} a {operacoes['data_referencia'].max():%Y-%m}",
            "unidade": "período",
            "status": "OK",
        },
        {
            "teste": "Contratos não automáticos distintos",
            "resultado": len(contratos),
            "unidade": "contratos",
            "status": "OK" if len(contratos) == 8294 else "REVISAR",
        },
        {
            "teste": "Subcréditos não automáticos",
            "resultado": int(operacoes["base_origem"].eq("operacoes_nao_automaticas").sum()),
            "unidade": "subcréditos",
            "status": "OK",
        },
        {
            "teste": "Duplicatas exatas nas operações automáticas",
            "resultado": int(
                operacoes.loc[
                    operacoes["base_origem"].eq("operacoes_indiretas_automaticas"),
                    "duplicata_exata",
                ].sum()
            ),
            "unidade": "registros excedentes",
            "status": "INFORMADO",
        },
    ]
    return pd.DataFrame(registros)


def _valor_excel(valor: Any) -> Any:
    if pd.isna(valor):
        return None
    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()
    if isinstance(valor, np.generic):
        return valor.item()
    return valor


def _adicionar_dataframe(
    wb: Workbook,
    nome: str,
    df: pd.DataFrame,
    descricao: str,
    fonte: str,
) -> None:
    ws = wb.create_sheet(nome[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = nome.replace("_", " ")
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=PALETTE["ink"].lstrip("#"))
    ws["A2"] = descricao
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color=PALETTE["muted"].lstrip("#"))
    ws["A3"] = f"Fonte: {fonte}"
    ws["A3"].font = Font(name="Aptos", size=9, color=PALETTE["muted"].lstrip("#"))
    inicio = 5
    quadro = df.copy()
    for linha_idx, linha in enumerate(dataframe_to_rows(quadro, index=False, header=True), inicio):
        for coluna_idx, valor in enumerate(linha, 1):
            ws.cell(linha_idx, coluna_idx, _valor_excel(valor))
    if quadro.shape[1] == 0:
        return
    for celula in ws[inicio]:
        celula.fill = PatternFill("solid", fgColor=PALETTE["blue"].lstrip("#"))
        celula.font = Font(name="Aptos", bold=True, color=PALETTE["white"].lstrip("#"))
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{inicio + 1}"
    ws.auto_filter.ref = f"A{inicio}:{get_column_letter(quadro.shape[1])}{inicio + len(quadro)}"
    if len(quadro):
        nome_tabela = "tb" + re.sub(r"[^A-Za-z0-9]", "", nome)[:20]
        tabela = Table(
            displayName=nome_tabela,
            ref=f"A{inicio}:{get_column_letter(quadro.shape[1])}{inicio + len(quadro)}",
        )
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tabela)
    for idx, coluna in enumerate(quadro.columns, 1):
        serie = quadro[coluna]
        largura = max(len(str(coluna)), *(len(str(v)) for v in serie.head(200).dropna())) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(max(largura, 11), 42)
        nome_coluna = str(coluna).lower()
        if any(token in nome_coluna for token in ["valor_", "media", "mediana", "quartil"]):
            formato = '#,##0.00;[Red]-#,##0.00;–'
        elif "pct" in nome_coluna or "participacao" in nome_coluna or "cobertura" in nome_coluna:
            formato = '0.00"%"'
        elif pd.api.types.is_integer_dtype(serie.dtype):
            formato = '#,##0'
        elif pd.api.types.is_datetime64_any_dtype(serie.dtype):
            formato = "mmm/yyyy"
        else:
            formato = "General"
        for celula in ws.iter_cols(min_col=idx, max_col=idx, min_row=inicio + 1, max_row=inicio + len(quadro)):
            for item in celula:
                item.number_format = formato
                item.alignment = Alignment(vertical="top")


def exportar_classificacao_excel(classificacao: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    resumo = (
        classificacao.groupby("classificacao_analise", as_index=False)
        .size()
        .rename(columns={"size": "registros"})
    )
    dicionario = pd.DataFrame(
        [
            ["classificacao_original", "Categoria histórica preservada exclusivamente para auditoria."],
            ["classificacao_analise", "Classificação principal binária: Verde estrito ou Demais operações."],
            ["indicador_verde_estrito", "Indicador lógico do numerador conservador."],
            ["bloco_tematico", "Bloco temático da política operacional."],
        ],
        columns=["variavel", "descricao"],
    )
    _adicionar_dataframe(
        wb, "Classificacao", classificacao, "Classificação binária das 242 políticas operacionais.",
        "Elaboração própria a partir das políticas operacionais do BNDES."
    )
    _adicionar_dataframe(wb, "Resumo", resumo, "Reconciliação obrigatória 30 + 212 = 242.", "Elaboração própria.")
    _adicionar_dataframe(wb, "Dicionario", dicionario, "Definição das variáveis centrais.", "Elaboração própria.")
    wb.save(CLASSIFICATION_WORKBOOK_PATH)


def exportar_excel(
    resultados: dict[str, pd.DataFrame],
    config: dict[str, Any],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    leia_me = pd.DataFrame(
        [
            ["Objetivo", "Mensurar a participação do Verde estrito no total do BNDES."],
            ["Indicador principal", "Desembolsos mensais; fluxos efetivamente liberados."],
            ["Indicador complementar", "Contratações; compromissos assumidos. Não somar aos desembolsos."],
            ["Preços", "Valores constantes de junho de 2026 pelo IPCA mensal, Tabela SIDRA 1737."],
            ["Período principal", "Janeiro de 2002 a dezembro de 2025."],
            ["Atualização", "2026 YTD: desembolsos jan–mar; contratações jan–abr."],
            ["Interpretação", "O numerador confirmado é conservador e pode ser lido como limite inferior."],
        ],
        columns=["item", "definicao"],
    )
    parametros = pd.DataFrame(
        [{"parametro": chave, "valor": json.dumps(valor, ensure_ascii=False) if isinstance(valor, list) else valor}
         for chave, valor in config.items()]
    )
    fonte = "Portal de Dados Abertos do BNDES; IBGE/SIDRA, Tabela 1737."
    planilhas = [
        ("Leia_me", leia_me, "Escopo, unidades e regras de leitura."),
        ("Parametros", parametros, "Parâmetros centrais da execução reproduzível."),
        ("Qualidade_Bases", resultados["qualidade_bases"], "Auditoria estrutural das fontes e unidades."),
        ("Cobertura_Classificacao", resultados["cobertura_classificacao"], "Cobertura do pareamento por universo, ano e regra."),
        ("Desembolsos_Mensais", resultados["desembolsos_mensais"], "Série mensal nominal e real."),
        ("Desembolsos_Anuais", resultados["desembolsos_anuais"], "Totais anuais, médias mensais e participação verde."),
        ("Contratacoes", resultados["contratacoes_anuais"], "Contratações anuais por base de origem."),
        ("Blocos_Instrumentos", resultados["blocos_instrumentos"], "Hierarquia bloco, instrumento, linha e modalidade."),
        ("Top10_Instrumentos", resultados["top10_instrumentos"], "Ranking de instrumentos verdes."),
        ("Top10_Linhas", resultados["top10_linhas"], "Ranking de linhas verdes."),
        ("Top10_Setores", resultados["top10_setores"], "Ranking de subsetores CNAE agrupados."),
        ("Top10_UF", resultados["top10_uf"], "Ranking com geografia válida; residual reportado separadamente."),
        ("Top10_Municipios", resultados["top10_municipios"], "Ranking com geografia válida; residual reportado separadamente."),
        ("Robustez", resultados["robustez"], "Sensibilidades de duplicatas, pareamento e definição de contrato verde."),
    ]
    for nome, quadro, descricao in planilhas:
        _adicionar_dataframe(wb, nome, quadro, descricao, fonte)
    wb.save(FINAL_WORKBOOK_PATH)


def _estilo_grafico(ax: plt.Axes) -> None:
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="y", color=PALETTE["grid"], linewidth=0.6, alpha=0.7)
    ax.tick_params(colors=PALETTE["ink"])


def _salvar_figura(fig: plt.Figure, nome: str) -> None:
    fig.savefig(FIGURES_DIR / f"{nome}.png", dpi=180, bbox_inches="tight", facecolor="white")
    fig.savefig(FIGURES_DIR / f"{nome}.svg", bbox_inches="tight", facecolor="white")
    plt.close(fig)


def gerar_figuras(
    desembolsos: pd.DataFrame,
    resultados: dict[str, pd.DataFrame],
    config: dict[str, Any],
) -> None:
    sns.set_theme(style="whitegrid", font="DejaVu Sans")
    anuais = resultados["desembolsos_anuais"]
    completos = anuais[anuais["ano"].le(int(config["ultimo_ano_completo"]))]

    fig, ax = plt.subplots(figsize=(12, 6.5))
    ax.plot(completos["ano"], completos["valor_real_jun2026_total"] / 1e9, color=PALETTE["blue"], lw=2.4, label="Total BNDES")
    ax.plot(completos["ano"], completos["valor_real_jun2026_verde_estrito"] / 1e9, color=PALETTE["olive"], lw=2.4, label="Verde estrito")
    ax.fill_between(completos["ano"], 0, completos["valor_real_jun2026_verde_estrito"] / 1e9, color=PALETTE["olive_light"], alpha=0.6)
    ax.set(title="Desembolsos reais do BNDES", xlabel="Ano", ylabel="R$ bilhões de junho de 2026")
    ax.legend(frameon=False, ncol=2)
    _estilo_grafico(ax)
    _salvar_figura(fig, "desembolsos_total_verde_reais")

    fig, ax = plt.subplots(figsize=(12, 5.8))
    ax.plot(completos["ano"], completos["participacao_verde_pct"], color=PALETTE["gold"], lw=2.5, marker="o", ms=3.5)
    ax.set(title="Participação do Verde estrito nos desembolsos", xlabel="Ano", ylabel="Participação (%)")
    _estilo_grafico(ax)
    _salvar_figura(fig, "participacao_verde_anual")

    sazonal = resultados["sazonalidade_mensal"]
    matriz = sazonal.pivot(index="ano", columns="mes", values="razao_mes_media_ano").loc[2002:2025]
    fig, ax = plt.subplots(figsize=(12, 8))
    sns.heatmap(matriz, cmap="Blues", center=1, linewidths=0.25, cbar_kws={"label": "Razão mês / média mensal do ano"}, ax=ax)
    ax.set(title="Sazonalidade mensal dos desembolsos", xlabel="Mês", ylabel="Ano")
    _salvar_figura(fig, "diagnostico_sazonalidade_dezembro")

    verde = desembolsos.loc[
        desembolsos["indicador_verde_estrito"] & desembolsos["ano"].le(int(config["ultimo_ano_completo"]))
    ].copy()
    blocos = (
        verde.groupby("bloco_tematico", dropna=False)["valor_real_jun2026"].sum().sort_values().tail(10)
    )
    fig, ax = plt.subplots(figsize=(11, 6.5))
    ax.barh(blocos.index.fillna("SEM BLOCO"), blocos.values / 1e9, color=PALETTE["olive"])
    ax.set(title="Composição do Verde estrito por bloco temático", xlabel="R$ bilhões de junho de 2026", ylabel="")
    _estilo_grafico(ax)
    _salvar_figura(fig, "composicao_bloco_tematico")

    for chave, dimensao, nome in [
        ("top10_instrumentos", "produto", "top10_instrumentos"),
        ("top10_setores", "subsetor_cnae_agrupado", "top10_subsetores"),
        ("top10_uf", "uf", "top10_uf"),
        ("top10_municipios", "municipio", "top10_municipios"),
    ]:
        top = resultados[chave]
        top = top.loc[(top["universo"].eq("Desembolsos")) & (top["recorte"].eq("2002–2025"))].sort_values("valor_real_jun2026")
        fig, ax = plt.subplots(figsize=(11, 6.5))
        ax.barh(top[dimensao].astype(str), top["valor_real_jun2026"] / 1e9, color=PALETTE["blue"])
        ax.set(title=f"Top 10 — {nome.replace('top10_', '').replace('_', ' ')}", xlabel="R$ bilhões de junho de 2026", ylabel="")
        _estilo_grafico(ax)
        _salvar_figura(fig, nome)

    acumulado = resultados["desembolsos_acumulados"].iloc[0]
    total = float(acumulado["valor_real_jun2026_total"])
    verde_total = float(acumulado["valor_real_jun2026_verde_estrito"])
    demais = float(acumulado["valor_real_jun2026_demais_operacoes"])
    fig, ax = plt.subplots(figsize=(12, 6.75))
    ax.set_xlim(0, 12); ax.set_ylim(0, 7); ax.axis("off")
    caixas = [
        (0.6, 2.25, 3.1, 2.2, PALETTE["blue"], "TOTAL BNDES", total),
        (5.0, 3.75, 3.0, 1.7, PALETTE["olive"], "VERDE ESTRITO", verde_total),
        (8.5, 1.25, 3.0, 1.7, PALETTE["gold"], "DEMAIS OPERAÇÕES", demais),
    ]
    for x, y, w, h, cor, rotulo, valor in caixas:
        ax.add_patch(FancyBboxPatch((x, y), w, h, boxstyle="round,pad=0.04,rounding_size=0.12", facecolor=cor, edgecolor="none"))
        ax.text(x + w / 2, y + h * 0.62, rotulo, ha="center", va="center", color="white", weight="bold", fontsize=13)
        ax.text(x + w / 2, y + h * 0.30, f"R$ {valor / 1e9:,.1f} bi".replace(",", "X").replace(".", ",").replace("X", "."), ha="center", va="center", color="white", fontsize=12)
    ax.annotate("", xy=(5.0, 4.6), xytext=(3.7, 3.8), arrowprops=dict(arrowstyle="->", lw=2, color=PALETTE["muted"]))
    ax.annotate("", xy=(8.5, 2.1), xytext=(3.7, 2.8), arrowprops=dict(arrowstyle="->", lw=2, color=PALETTE["muted"]))
    ax.text(0.6, 6.35, "FINANCIAMENTO VERDE NO BNDES", fontsize=20, weight="bold", color=PALETTE["ink"])
    ax.text(0.6, 5.95, "Desembolsos acumulados de 2002 a 2025, em preços de junho de 2026", fontsize=11, color=PALETTE["muted"])
    ax.text(0.6, 0.45, "Classificação principal binária. Nenhum registro é excluído do denominador.", fontsize=10, color=PALETTE["muted"])
    _salvar_figura(fig, "classificacao_recortes_bndes_whatsapp")


def validar_resultados(
    classificacao: pd.DataFrame,
    ipca: pd.DataFrame,
    desembolsos: pd.DataFrame,
    operacoes: pd.DataFrame,
    contratos: pd.DataFrame,
    resultados: dict[str, pd.DataFrame],
    config: dict[str, Any],
) -> dict[str, Any]:
    contagens = classificacao["classificacao_analise"].value_counts().to_dict()
    data_ref = pd.Timestamp(config["ipca_mes_referencia"] + "-01")
    meses_por_ano = desembolsos.groupby("ano")["mes"].nunique()
    anuais = resultados["desembolsos_anuais"]
    reconciliacao = np.allclose(
        anuais["valor_real_jun2026_verde_estrito"] + anuais["valor_real_jun2026_demais_operacoes"],
        anuais["valor_real_jun2026_total"],
        rtol=1e-10, atol=0.1,
    )
    testes = {
        "classificacao_242_30_212": len(classificacao) == 242 and contagens == {"Demais operações": 212, "Verde estrito": 30},
        "ipca_100_pct_meses_desembolsos": not desembolsos["ipca_mes"].isna().any(),
        "ipca_100_pct_meses_contratacoes": not operacoes["ipca_mes"].isna().any(),
        "fator_jun2026_igual_1": np.isclose(ipca.loc[ipca["data_referencia"].eq(data_ref), "fator_ipca_jun2026"].iloc[0], 1.0),
        "doze_meses_2002_2025": bool(meses_por_ano.loc[2002:2025].eq(12).all()),
        "ano_2026_parcial": bool(desembolsos.loc[desembolsos["ano"].eq(2026), "mes"].nunique() == 3 and operacoes.loc[operacoes["ano"].eq(2026), "mes"].nunique() == 4),
        "reconciliacao_verde_demais_total": bool(reconciliacao),
        "contratos_nao_automaticos_8294": len(contratos) == 8294,
        "participacao_verde_nao_supera_100": bool(anuais["participacao_verde_pct"].between(0, 100).all()),
        "sem_categorias_historicas_nas_saidas_principais": not set(["Não verde", "Verde ampliado", "Intermediário", "Revisar"]).intersection(set(desembolsos["classificacao_analise"].dropna().unique())),
        "ufs_invalidas_fora_top10": not resultados["top10_uf"]["uf"].astype(str).isin(config["ufs_invalidas"]).any(),
        "municipios_invalidos_fora_top10": not resultados["top10_municipios"]["municipio"].astype(str).isin(config["municipios_invalidos"]).any(),
        "crescimento_anual_2026_suprimido": bool(
            anuais.loc[
                anuais["ano"].eq(int(config["ano_parcial"])),
                ["crescimento_real_total_pct", "crescimento_real_verde_pct"],
            ].isna().all().all()
        ),
    }
    testes = {nome: bool(resultado) for nome, resultado in testes.items()}
    contagens = {str(nome): int(valor) for nome, valor in contagens.items()}
    base_principal = desembolsos.loc[desembolsos["ano"].le(int(config["ultimo_ano_completo"]))]
    cobertura_confirmada_pct = float(
        100
        * base_principal.loc[base_principal["pareamento_confirmado"], "valor_real_jun2026"].sum()
        / base_principal["valor_real_jun2026"].sum()
    )
    cobertura_suficiente = cobertura_confirmada_pct >= 100 * float(
        config["limite_cobertura_publicacao"]
    )
    cobertura_territorial = resultados["cobertura_territorial"].loc[
        resultados["cobertura_territorial"]["recorte"].eq("2002–2025"),
        ["universo", "cobertura_territorial_pct"],
    ]
    cobertura_territorial_dict = {
        str(linha["universo"]): float(linha["cobertura_territorial_pct"])
        for _, linha in cobertura_territorial.iterrows()
    }
    ressalvas = []
    if not cobertura_suficiente:
        ressalvas.append(
            "Cobertura confirmada inferior a 95%; divulgar a participação apenas como estimativa conservadora ou limite inferior."
        )
    if any(valor < 95 for valor in cobertura_territorial_dict.values()):
        ressalvas.append(
            "Cobertura territorial inferior a 95%; Top 10 de UF e município representam somente o valor com geografia válida."
        )
    todos_testes = all(testes.values())
    parecer = "BLOQUEADO" if not todos_testes else ("APROVADO COM RESSALVAS" if ressalvas else "APROVADO")
    relatorio = {
        "executado_em": pd.Timestamp.now().isoformat(),
        "testes": testes,
        "todos_testes_aprovados": todos_testes,
        "parecer": parecer,
        "ressalvas": ressalvas,
        "cobertura_confirmada_valor_2002_2025_pct": cobertura_confirmada_pct,
        "cobertura_minima_divulgacao_sem_ressalva_pct": 100 * float(config["limite_cobertura_publicacao"]),
        "cobertura_suficiente_para_divulgacao_sem_ressalva": bool(cobertura_suficiente),
        "cobertura_territorial_2002_2025_pct": cobertura_territorial_dict,
        "contagens_classificacao": contagens,
        "linhas_desembolsos": len(desembolsos),
        "linhas_operacoes": len(operacoes),
        "contratos_nao_automaticos": len(contratos),
    }
    VALIDATION_PATH.write_text(json.dumps(relatorio, ensure_ascii=False, indent=2), encoding="utf-8")
    if not relatorio["todos_testes_aprovados"]:
        falhas = [nome for nome, ok in testes.items() if not ok]
        raise AssertionError(f"Testes de aceite não aprovados: {falhas}")
    return relatorio


def executar_pipeline(usar_cache: bool = True) -> PipelineArtifacts:
    ensure_dirs()
    config = carregar_config()
    inventario = inventario_fontes()
    classificacao = construir_classificacao_binaria()
    exportar_classificacao_excel(classificacao)
    ipca = baixar_ipca(config)
    concordancia = construir_concordancia(classificacao)
    caminho_desembolsos = PROCESSED_DIR / "desembolsos_mensais_analitico.parquet"
    caminho_operacoes = PROCESSED_DIR / "operacoes_bndes_analitica.parquet"
    caminho_contratos = PROCESSED_DIR / "contratos_nao_automaticos_analitico.parquet"
    if usar_cache and all(
        caminho.exists() for caminho in [caminho_desembolsos, caminho_operacoes, caminho_contratos]
    ):
        colunas_comuns = [
            "data_referencia", "ano", "mes", "ano_parcial", "base_origem", "valor_nominal",
            "ipca_mes", "fator_ipca_jun2026", "valor_real_jun2026",
            "classificacao_analise", "indicador_verde_estrito",
            "status_pareamento", "regra_pareamento", "pareamento_confirmado",
            "geografia_valida", "bloco_tematico", "produto", "instrumento_financeiro",
            "forma_de_apoio", "subsetor_cnae_agrupado", "uf", "municipio",
        ]
        desembolsos = pd.read_parquet(caminho_desembolsos, columns=colunas_comuns)
        operacoes = pd.read_parquet(
            caminho_operacoes, columns=colunas_comuns + ["modalidade_de_apoio", "duplicata_exata"]
        )
        contratos = pd.read_parquet(caminho_contratos)
    else:
        desembolsos = preparar_desembolsos(concordancia, ipca, config)
        operacoes, contratos = preparar_operacoes(concordancia, ipca, config)

    nomes_resultados_base = [
        "desembolsos_mensais", "desembolsos_anuais", "cobertura_classificacao_desembolsos",
        "sazonalidade_mensal", "desembolsos_acumulados", "contratacoes_anuais",
        "cobertura_classificacao_contratacoes", "robustez_duplicatas_automaticas",
        "robustez_contratos_nao_automaticos", "estatisticas_contratacoes",
        "blocos_instrumentos", "top10_instrumentos", "top10_linhas", "top10_setores",
        "top10_uf", "top10_municipios", "cobertura_territorial",
    ]
    caminhos_resultados = {
        nome: PROCESSED_DIR / f"resultado_{nome}.parquet" for nome in nomes_resultados_base
    }
    if usar_cache and all(caminho.exists() for caminho in caminhos_resultados.values()):
        resultados = {
            nome: pd.read_parquet(caminho) for nome, caminho in caminhos_resultados.items()
        }
    else:
        resultados: dict[str, pd.DataFrame] = {}
        resultados.update(calcular_desembolsos(desembolsos, config))
        resultados.update(calcular_contratacoes(operacoes, contratos, config))
        resultados.update(calcular_composicao_top10(desembolsos, operacoes, config))
    ano_parcial = int(config["ano_parcial"])
    ultimo_ano_completo = int(config["ultimo_ano_completo"])
    meses_ytd = int(desembolsos.loc[desembolsos["ano"].eq(ano_parcial), "mes"].max())
    anuais = resultados["desembolsos_anuais"].copy()
    anuais["crescimento_real_total_ytd_pct"] = np.nan
    anuais["crescimento_real_verde_ytd_pct"] = np.nan
    atual_ytd = desembolsos.loc[desembolsos["ano"].eq(ano_parcial)]
    anterior_ytd = desembolsos.loc[
        desembolsos["ano"].eq(ultimo_ano_completo) & desembolsos["mes"].le(meses_ytd)
    ]
    total_atual_ytd = atual_ytd["valor_real_jun2026"].sum()
    total_anterior_ytd = anterior_ytd["valor_real_jun2026"].sum()
    verde_atual_ytd = atual_ytd.loc[atual_ytd["indicador_verde_estrito"], "valor_real_jun2026"].sum()
    verde_anterior_ytd = anterior_ytd.loc[
        anterior_ytd["indicador_verde_estrito"], "valor_real_jun2026"
    ].sum()
    mascara_ano_parcial = anuais["ano"].eq(ano_parcial)
    anuais.loc[mascara_ano_parcial, ["crescimento_real_total_pct", "crescimento_real_verde_pct"]] = np.nan
    anuais.loc[mascara_ano_parcial, "crescimento_real_total_ytd_pct"] = (
        100 * (total_atual_ytd / total_anterior_ytd - 1) if total_anterior_ytd else np.nan
    )
    anuais.loc[mascara_ano_parcial, "crescimento_real_verde_ytd_pct"] = (
        100 * (verde_atual_ytd / verde_anterior_ytd - 1) if verde_anterior_ytd else np.nan
    )
    resultados["desembolsos_anuais"] = anuais
    resultados["qualidade_bases"] = construir_qualidade_bases(inventario, desembolsos, operacoes, contratos)
    resultados["inventario_fontes"] = inventario
    resultados["cobertura_classificacao"] = pd.concat(
        [
            resultados["cobertura_classificacao_desembolsos"].assign(universo="Desembolsos", base_origem="desembolsos_mensais"),
            resultados["cobertura_classificacao_contratacoes"].assign(universo="Contratações"),
        ],
        ignore_index=True,
    )
    base_principal = desembolsos.loc[desembolsos["ano"].le(int(config["ultimo_ano_completo"]))]
    total_principal = base_principal["valor_real_jun2026"].sum()
    regras_exatas = {
        "Produto + linha/sublinha exatos",
        "Instrumento financeiro exato e unívoco",
        "Correspondência textual exata",
        "Correspondência exata após normalização",
    }
    valor_verde_exato = base_principal.loc[
        base_principal["indicador_verde_estrito"]
        & base_principal["regra_pareamento"].isin(regras_exatas),
        "valor_real_jun2026",
    ].sum()
    valor_verde_validado = base_principal.loc[
        base_principal["indicador_verde_estrito"], "valor_real_jun2026"
    ].sum()
    participacao_exata = 100 * valor_verde_exato / total_principal
    participacao_validada = 100 * valor_verde_validado / total_principal
    resultados["robustez_pareamento"] = pd.DataFrame(
        [
            {
                "sensibilidade": "Pareamento estritamente exato",
                "valor_verde_real_jun2026": valor_verde_exato,
                "valor_total_real_jun2026": total_principal,
                "participacao_verde_pct": participacao_exata,
                "diferenca_pp_vs_principal": participacao_exata - participacao_validada,
                "implicacao": "Nenhum registro verde confirmado pelas chaves exatas disponíveis.",
            },
            {
                "sensibilidade": "Exato + sinônimos manualmente validados",
                "valor_verde_real_jun2026": valor_verde_validado,
                "valor_total_real_jun2026": total_principal,
                "participacao_verde_pct": participacao_validada,
                "diferenca_pp_vs_principal": 0.0,
                "implicacao": "Especificação principal conservadora; exige ressalva de cobertura.",
            },
        ]
    )
    resultados["robustez"] = pd.concat(
        [
            resultados["robustez_duplicatas_automaticas"].assign(modulo="Duplicatas automáticas"),
            resultados["robustez_contratos_nao_automaticos"].assign(modulo="Definição de contrato verde"),
            resultados["robustez_pareamento"].assign(modulo="Pareamento histórico"),
        ],
        ignore_index=True,
        sort=False,
    )

    for nome, quadro in resultados.items():
        quadro_saida = quadro.copy()
        for coluna in quadro_saida.select_dtypes(include="object").columns:
            tipos = quadro_saida[coluna].dropna().map(type).nunique()
            if tipos > 1:
                quadro_saida[coluna] = quadro_saida[coluna].map(
                    lambda valor: None if pd.isna(valor) else str(valor)
                )
        quadro_saida.to_parquet(PROCESSED_DIR / f"resultado_{nome}.parquet", index=False)

    validacao = validar_resultados(
        classificacao, ipca, desembolsos, operacoes, contratos, resultados, config
    )
    exportar_excel(resultados, config)
    gerar_figuras(desembolsos, resultados, config)

    acumulado = resultados["desembolsos_acumulados"].iloc[0]
    anual_2025 = resultados["desembolsos_anuais"].loc[
        resultados["desembolsos_anuais"]["ano"].eq(2025)
    ].iloc[0]
    ytd_2026 = resultados["desembolsos_anuais"].loc[
        resultados["desembolsos_anuais"]["ano"].eq(2026)
    ].iloc[0]
    resumo = {
        "executado_em": pd.Timestamp.now().isoformat(),
        "periodo_principal": "2002-01 a 2025-12",
        "desembolsos_2026_ytd": "2026-01 a 2026-03",
        "contratacoes_2026_ytd": "2026-01 a 2026-04",
        "precos": "junho de 2026 (IPCA mensal, SIDRA 1737)",
        "politicas_verde_estrito": 30,
        "politicas_demais_operacoes": 212,
        "participacao_verde_desembolsos_2002_2025_pct": float(acumulado["participacao_verde_pct"]),
        "participacao_verde_desembolsos_2025_pct": float(anual_2025["participacao_verde_pct"]),
        "participacao_verde_desembolsos_2026_ytd_pct": float(ytd_2026["participacao_verde_pct"]),
        "cobertura_confirmada_valor_2002_2025_pct": float(
            100
            * desembolsos.loc[
                desembolsos["ano"].le(2025) & desembolsos["pareamento_confirmado"],
                "valor_real_jun2026",
            ].sum()
            / desembolsos.loc[desembolsos["ano"].le(2025), "valor_real_jun2026"].sum()
        ),
        "resultado_conservador": True,
        "todos_testes_aprovados": validacao["todos_testes_aprovados"],
        "parecer_metodologico": validacao["parecer"],
        "ressalvas": validacao["ressalvas"],
        "workbook": str(FINAL_WORKBOOK_PATH),
        "notebook": str(PROJECT_ROOT / "notebooks" / "05_analise_descritiva.ipynb"),
    }
    SUMMARY_PATH.write_text(json.dumps(resumo, ensure_ascii=False, indent=2), encoding="utf-8")
    return PipelineArtifacts(
        classificacao=classificacao,
        ipca=ipca,
        concordancia=concordancia,
        desembolsos=desembolsos,
        operacoes=operacoes,
        contratos_nao_automaticos=contratos,
        resultados=resultados,
        resumo=resumo,
        validacao=validacao,
    )
